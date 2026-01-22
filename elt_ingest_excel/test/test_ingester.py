"""Tests for Excel ingestion into DuckDB."""

import tempfile
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook

from elt_ingest_excel import (
    ExcelIngester,
    ExcelIngestConfig,
    WorkbookConfig,
    SheetConfig,
    JsonConfigParser,
    ExcelLoader,
)


@pytest.fixture
def sample_workbook():
    """Create a sample Excel workbook for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        workbook_path = f.name

    wb = Workbook()

    # Sheet1 - simple data
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Age"
    ws1["C1"] = "City"
    ws1["A2"] = "Alice"
    ws1["B2"] = 30
    ws1["C2"] = "New York"
    ws1["A3"] = "Bob"
    ws1["B3"] = 25
    ws1["C3"] = "Los Angeles"
    ws1["A4"] = "Charlie"
    ws1["B4"] = 35
    ws1["C4"] = "Chicago"

    # Sheet2 - different data
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Product"
    ws2["B1"] = "Price"
    ws2["C1"] = "Quantity"
    ws2["A2"] = "Widget"
    ws2["B2"] = 9.99
    ws2["C2"] = 100
    ws2["A3"] = "Gadget"
    ws2["B3"] = 19.99
    ws2["C3"] = 50

    # Sheet3 - with header on row 3
    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "Report Title"
    ws3["A2"] = "Generated: 2024-01-01"
    ws3["A3"] = "ID"
    ws3["B3"] = "Value"
    ws3["A4"] = 1
    ws3["B4"] = 100
    ws3["A5"] = 2
    ws3["B5"] = 200

    wb.save(workbook_path)
    wb.close()

    yield workbook_path

    # Cleanup
    Path(workbook_path).unlink(missing_ok=True)


@pytest.fixture
def temp_database():
    """Create a temporary DuckDB database path."""
    # Use a temp directory and generate a path (don't create the file)
    temp_dir = tempfile.mkdtemp()
    db_path = Path(temp_dir) / "test.duckdb"

    yield str(db_path)

    # Cleanup
    db_path.unlink(missing_ok=True)
    Path(temp_dir).rmdir()


class TestExcelLoader:
    """Tests for ExcelLoader."""

    def test_load_simple_sheet(self, sample_workbook):
        """Test loading a simple sheet."""
        sheet_config = SheetConfig(
            sheet_name="Sheet1",
            target_table_name="test_table",
        )

        with ExcelLoader(sample_workbook) as loader:
            data = loader.load_sheet(sheet_config)

        assert len(data) == 3
        assert data[0]["name"] == "Alice"
        assert data[0]["age"] == 30
        assert data[0]["city"] == "New York"
        assert data[2]["name"] == "Charlie"

    def test_load_sheet_with_custom_header(self, sample_workbook):
        """Test loading a sheet with custom header row."""
        sheet_config = SheetConfig(
            sheet_name="Sheet3",
            target_table_name="test_table",
            header_row=3,
            skip_rows=0,
        )

        with ExcelLoader(sample_workbook) as loader:
            data = loader.load_sheet(sheet_config)

        assert len(data) == 2
        assert data[0]["id"] == 1
        assert data[0]["value"] == 100
        assert data[1]["id"] == 2
        assert data[1]["value"] == 200

    def test_get_sheet_names(self, sample_workbook):
        """Test getting sheet names from workbook."""
        with ExcelLoader(sample_workbook) as loader:
            names = loader.get_sheet_names()

        assert "Sheet1" in names
        assert "Sheet2" in names
        assert "Sheet3" in names

    def test_sheet_not_found(self, sample_workbook):
        """Test error when sheet doesn't exist."""
        sheet_config = SheetConfig(
            sheet_name="NonexistentSheet",
            target_table_name="test_table",
        )

        with pytest.raises(ValueError, match="Sheet 'NonexistentSheet' not found"):
            with ExcelLoader(sample_workbook) as loader:
                loader.load_sheet(sheet_config)

    def test_workbook_not_found(self):
        """Test error when workbook doesn't exist."""
        with pytest.raises(FileNotFoundError):
            ExcelLoader("/nonexistent/workbook.xlsx")


class TestExcelIngester:
    """Tests for ExcelIngester."""

    def test_ingest_single_sheet(self, sample_workbook, temp_database):
        """Test ingesting a single sheet."""
        config = ExcelIngestConfig(
            database_path=Path(temp_database),
            workbooks=[
                WorkbookConfig(
                    workbook_file_name=sample_workbook,
                    sheets=[
                        SheetConfig(
                            sheet_name="Sheet1",
                            target_table_name="people",
                        )
                    ],
                )
            ],
        )

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()

        assert len(results) == 1
        assert results[0].table_name == "people"
        assert results[0].rows_loaded == 3
        assert results[0].row_count == 3

        # Verify data in DuckDB
        conn = duckdb.connect(temp_database)
        data = conn.execute("SELECT * FROM people ORDER BY name").fetchall()
        conn.close()

        assert len(data) == 3
        assert data[0][0] == "Alice"  # name column

    def test_ingest_multiple_sheets(self, sample_workbook, temp_database):
        """Test ingesting multiple sheets from one workbook."""
        config = ExcelIngestConfig(
            database_path=Path(temp_database),
            workbooks=[
                WorkbookConfig(
                    workbook_file_name=sample_workbook,
                    sheets=[
                        SheetConfig(
                            sheet_name="Sheet1",
                            target_table_name="people",
                        ),
                        SheetConfig(
                            sheet_name="Sheet2",
                            target_table_name="products",
                        ),
                    ],
                )
            ],
        )

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()

        assert len(results) == 2
        assert results[0].table_name == "people"
        assert results[0].row_count == 3
        assert results[1].table_name == "products"
        assert results[1].row_count == 2

    def test_ingest_from_json_config(self, sample_workbook, temp_database):
        """Test ingesting using JSON configuration."""
        config_data = [
            {
                "workbookFileName": sample_workbook,
                "sheets": [
                    {"sheetName": "Sheet1", "targetTableName": "people"},
                    {"sheetName": "Sheet2", "targetTableName": "products"},
                ],
            }
        ]

        config = JsonConfigParser.from_json(
            config_data,
            database_path=temp_database,
        )

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()

        assert len(results) == 2

        # Verify both tables exist and have correct data
        conn = duckdb.connect(temp_database)
        people_count = conn.execute("SELECT COUNT(*) FROM people").fetchone()[0]
        products_count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        conn.close()

        assert people_count == 3
        assert products_count == 2

    def test_ingest_with_custom_header_row(self, sample_workbook, temp_database):
        """Test ingesting sheet with custom header row."""
        config = ExcelIngestConfig(
            database_path=Path(temp_database),
            workbooks=[
                WorkbookConfig(
                    workbook_file_name=sample_workbook,
                    sheets=[
                        SheetConfig(
                            sheet_name="Sheet3",
                            target_table_name="report_data",
                            header_row=3,
                        ),
                    ],
                )
            ],
        )

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()

        assert results[0].row_count == 2

        # Verify data
        conn = duckdb.connect(temp_database)
        data = conn.execute("SELECT * FROM report_data ORDER BY id").fetchall()
        conn.close()

        assert len(data) == 2
        assert data[0][0] == 1
        assert data[0][1] == 100

    def test_replace_data(self, sample_workbook, temp_database):
        """Test that replace_data=True replaces existing data."""
        config = ExcelIngestConfig(
            database_path=Path(temp_database),
            workbooks=[
                WorkbookConfig(
                    workbook_file_name=sample_workbook,
                    sheets=[
                        SheetConfig(
                            sheet_name="Sheet1",
                            target_table_name="people",
                        )
                    ],
                )
            ],
            replace_data=True,
        )

        # Run twice
        with ExcelIngester(config) as ingester:
            ingester.ingest()

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()

        # Should still have only 3 rows (replaced, not appended)
        assert results[0].row_count == 3

    def test_print_summary(self, sample_workbook, temp_database, capsys):
        """Test that print_summary outputs correctly."""
        config = ExcelIngestConfig(
            database_path=Path(temp_database),
            workbooks=[
                WorkbookConfig(
                    workbook_file_name=sample_workbook,
                    sheets=[
                        SheetConfig(
                            sheet_name="Sheet1",
                            target_table_name="people",
                        )
                    ],
                )
            ],
        )

        with ExcelIngester(config) as ingester:
            results = ingester.ingest()
            ingester.print_summary(results)

        captured = capsys.readouterr()
        assert "INGESTION SUMMARY" in captured.out
        assert "people" in captured.out
        assert "Sheet1" in captured.out
        assert "Rows loaded: 3" in captured.out
        assert "Table count: 3" in captured.out
