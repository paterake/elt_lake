"""Excel file loader using openpyxl."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import SheetConfig


class ExcelLoader:
    """Loads data from Excel worksheets.

    Uses openpyxl to read Excel files and extract sheet data
    as a list of dictionaries suitable for loading into DuckDB.
    """

    def __init__(self, workbook_path: str | Path):
        """Initialize the loader with a workbook path.

        Args:
            workbook_path: Path to the Excel workbook file.

        Raises:
            FileNotFoundError: If the workbook file doesn't exist.
        """
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        self._workbook = None

    def __enter__(self):
        """Context manager entry - load the workbook."""
        self._workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - close the workbook."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def load_sheet(self, sheet_config: SheetConfig) -> list[dict[str, Any]]:
        """Load data from a worksheet.

        Args:
            sheet_config: Configuration for the sheet to load.

        Returns:
            List of dictionaries, one per row, with column headers as keys.

        Raises:
            ValueError: If the sheet doesn't exist in the workbook.
            RuntimeError: If the loader is not in a context manager.
        """
        if self._workbook is None:
            raise RuntimeError("ExcelLoader must be used as a context manager")

        if sheet_config.sheet_name not in self._workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_config.sheet_name}' not found in workbook. "
                f"Available sheets: {self._workbook.sheetnames}"
            )

        worksheet: Worksheet = self._workbook[sheet_config.sheet_name]
        return self._extract_data(worksheet, sheet_config)

    def _extract_data(
        self,
        worksheet: Worksheet,
        sheet_config: SheetConfig,
    ) -> list[dict[str, Any]]:
        """Extract data from a worksheet.

        Args:
            worksheet: The openpyxl worksheet object.
            sheet_config: Configuration for the sheet.

        Returns:
            List of dictionaries with column headers as keys.
        """
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            return []

        # Calculate the actual header row index (0-based from 1-indexed config)
        header_idx = sheet_config.header_row - 1

        if header_idx >= len(rows):
            return []

        # Extract headers
        headers = self._clean_headers(rows[header_idx])

        # Calculate data start index (0-based from 1-indexed config)
        data_start_idx = sheet_config.data_row - 1
        data = []

        for row in rows[data_start_idx:]:
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value

            data.append(row_dict)

        return data

    def _clean_headers(self, header_row: tuple) -> list[str]:
        """Clean and normalize column headers.

        Args:
            header_row: Tuple of header values from the worksheet.

        Returns:
            List of cleaned header strings.
        """
        headers = []
        for i, header in enumerate(header_row):
            if header is None:
                # Use a placeholder for empty headers
                headers.append(f"column_{i}")
            else:
                # Convert to string and clean
                header_str = str(header).strip()
                # Replace spaces and special chars with underscores
                header_str = self._normalize_column_name(header_str)
                headers.append(header_str)

        return headers

    def _normalize_column_name(self, name: str) -> str:
        """Normalize a column name for database compatibility.

        Args:
            name: Original column name.

        Returns:
            Normalized column name (lowercase, underscores for spaces).
        """
        # Replace spaces and common separators with underscores
        normalized = name.replace(" ", "_").replace("-", "_").replace(".", "_")
        # Convert to lowercase
        normalized = normalized.lower()
        # Remove any remaining special characters
        normalized = "".join(c if c.isalnum() or c == "_" else "" for c in normalized)
        # Ensure it doesn't start with a number
        if normalized and normalized[0].isdigit():
            normalized = f"col_{normalized}"
        return normalized

    def get_sheet_names(self) -> list[str]:
        """Get list of sheet names in the workbook.

        Returns:
            List of sheet names.

        Raises:
            RuntimeError: If the loader is not in a context manager.
        """
        if self._workbook is None:
            raise RuntimeError("ExcelLoader must be used as a context manager")
        return self._workbook.sheetnames
