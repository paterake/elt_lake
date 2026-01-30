"""Excel publisher using openpyxl for writing DuckDB data to Excel workbooks.

Note: openpyxl may not fully preserve drawing shapes in .xlsm files.
Use ExcelPublisherXlwings if you need to preserve all Excel features.
"""

from pathlib import Path
from typing import TYPE_CHECKING, Any, Union

from openpyxl import load_workbook

from ..models import PublishWorkbookConfig
from .base import BaseExcelPublisher, PublishResult

if TYPE_CHECKING:
    from ..reporting import PipelineReporter


class ExcelPublisherOpenpyxl(BaseExcelPublisher):
    """Publisher for writing DuckDB data to Excel workbooks using openpyxl.

    This class handles:
    1. Cloning a template workbook to target location
    2. Querying DuckDB tables for data
    3. Writing data to specific sheets in the workbook
    """

    def __init__(
        self,
        database_path: Union[str, Path],
        reporter: "PipelineReporter | None" = None,
    ):
        """Initialize the publisher.

        Args:
            database_path: Path to the DuckDB database file.
            reporter: Optional reporter for output messages.
        """
        super().__init__(database_path, reporter)

    def _open_and_process_workbook(
        self,
        workbook_config: PublishWorkbookConfig,
        tgt_path: Path,
    ) -> list[PublishResult]:
        """Open workbook with openpyxl and process all sheets.

        Args:
            workbook_config: Configuration for the workbook.
            tgt_path: Path to the target workbook file.

        Returns:
            List of PublishResult for each sheet processed.
        """
        # Open the target workbook
        wb = load_workbook(tgt_path, keep_vba=True)  # keep_vba for .xlsm files

        try:
            # Process each sheet
            results = self._process_sheets(wb, workbook_config)

            # Save the workbook
            self._report_saving(tgt_path)
            wb.save(tgt_path)
            self._report_saved(tgt_path)

        finally:
            wb.close()

        return results

    def _sheet_exists(self, workbook: Any, sheet_name: str) -> bool:
        """Check if a sheet exists in the workbook.

        Args:
            workbook: openpyxl Workbook object.
            sheet_name: Name of the sheet to check.

        Returns:
            True if sheet exists, False otherwise.
        """
        return sheet_name in workbook.sheetnames

    def _get_sheet(self, workbook: Any, sheet_name: str) -> Any:
        """Get a sheet from the workbook.

        Args:
            workbook: openpyxl Workbook object.
            sheet_name: Name of the sheet to get.

        Returns:
            openpyxl Worksheet object.
        """
        return workbook[sheet_name]

    def _write_data_to_sheet(
        self,
        sheet: Any,
        rows: list[tuple],
        start_row: int,
    ) -> None:
        """Write data to the sheet cell by cell.

        Args:
            sheet: openpyxl Worksheet object.
            rows: Data rows to write.
            start_row: Starting row number (1-indexed).
        """
        for row_idx, row_data in enumerate(rows):
            excel_row = start_row + row_idx
            for col_idx, value in enumerate(row_data):
                col = col_idx + 1  # Excel columns are 1-indexed
                sheet.cell(row=excel_row, column=col, value=value)
