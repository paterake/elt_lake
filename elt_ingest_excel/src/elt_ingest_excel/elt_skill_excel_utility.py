import argparse
import json
import os
import re
import zipfile
from typing import Iterable, List, Optional, Tuple


def _try_xlwings_filter(path: str, sheet: str, col: str, cell: Optional[str], row: Optional[int]) -> List[str]:
    try:
        import xlwings as xw
    except Exception:
        return []
    app = None
    vals = set()
    try:
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(path)
        sht = wb.sheets[sheet]
        targets: Iterable[int]
        target_rows: List[int] = []
        if cell:
            digits = "".join(ch for ch in cell if ch.isdigit())
            if digits.isdigit():
                target_rows.append(int(digits))
        elif row:
            target_rows.append(int(row))
        if not target_rows:
            used = sht.used_range
            last_row = used.last_cell.row
            end_row = min(last_row, 10000)
            target_rows = list(range(1, end_row + 1))
        targets = target_rows
        col_idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
        for r in targets:
            cell_rng = sht.range((r, col_idx))
            try:
                v = cell_rng.api.Validation
                t = v.Type
            except Exception:
                continue
            if t != 3:
                continue
            f = None
            try:
                f = v.Formula1
            except Exception:
                continue
            if not f:
                continue
            ref = f[1:] if isinstance(f, str) and f.startswith("=") else f
            if all(x not in ref for x in ("!", ":")):
                for part in str(ref).replace(";", ",").split(","):
                    p = part.strip()
                    if p:
                        vals.add(p)
                continue
            try:
                if "!" in ref:
                    sh, rng = ref.split("!", 1)
                    sh = sh.strip("'")
                    rng_obj = wb.sheets[sh].range(rng)
                else:
                    rng_obj = wb.names[ref].refers_to_range
                data = rng_obj.value
                if isinstance(data, list):
                    for rowv in data:
                        if isinstance(rowv, list):
                            for x in rowv:
                                if x not in (None, ""):
                                    vals.add(str(x).strip())
                        else:
                            if rowv not in (None, ""):
                                vals.add(str(rowv).strip())
                elif data not in (None, ""):
                    vals.add(str(data).strip())
            except Exception:
                continue
        if not vals:
            try:
                api = sht.api
                dds = api.DropDowns()
                for i in range(1, dds.Count + 1):
                    try:
                        dd = dds.Item(i)
                        tl = dd.TopLeftCell
                        col_match = tl.Column == (ord(col.upper()) - 64 if not col.isdigit() else int(col))
                        if not col_match:
                            continue
                        lfr = dd.ListFillRange
                        if not lfr:
                            continue
                        if "!" in lfr:
                            sh, rng = lfr.split("!", 1)
                            sh = sh.strip("'")
                            rng_obj = wb.sheets[sh].range(rng)
                        else:
                            rng_obj = wb.names[lfr].refers_to_range
                        data = rng_obj.value
                        if isinstance(data, list):
                            for rowv in data:
                                if isinstance(rowv, list):
                                    for x in rowv:
                                        if x not in (None, ""):
                                            vals.add(str(x).strip())
                                else:
                                    if rowv not in (None, ""):
                                        vals.add(str(rowv).strip())
                        elif data not in (None, ""):
                            vals.add(str(data).strip())
                    except Exception:
                        continue
            except Exception:
                pass
    except Exception:
        # Any macOS automation error should gracefully fall back
        return []
    finally:
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass
    return sorted(vals)


def _openpyxl_filter(path: str, sheet: str, col: str) -> List[str]:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries

    ws = load_workbook(path, data_only=True, keep_vba=True)[sheet]
    col_idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
    vals: List[str] = []
    dv_container = getattr(ws, "data_validations", None)
    if dv_container is not None:
        for dv in dv_container.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            applies = any(cr.min_col <= col_idx <= cr.max_col for cr in dv.ranges)
            if not applies:
                continue
            f1 = dv.formula1
            if not f1:
                continue
            f = f1[1:] if isinstance(f1, str) and f1.startswith("=") else f1
            if isinstance(f, str) and f.startswith('"') and f.endswith('"'):
                vals += [x.strip() for x in f.strip('"').split(",") if x.strip()]
                continue
            wb = ws.parent
            if isinstance(f, str) and "!" not in f and f in wb.defined_names:
                dn = wb.defined_names[f]
                for title, area in dn.destinations:
                    ws2 = wb[title]
                    min_c, min_r, max_c, max_r = range_boundaries(area)
                    for row in ws2.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
                        for cell in row:
                            if cell not in (None, ""):
                                vals.append(str(cell).strip())
                continue
            if isinstance(f, str) and "!" in f:
                if f.startswith("'"):
                    sh = f.split("'", 2)[1]
                    rng = f.split("'", 2)[2].lstrip("!")
                else:
                    sh, rng = f.split("!", 1)
                ws2 = wb[sh]
                min_c, min_r, max_c, max_r = range_boundaries(rng)
                for row in ws2.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
                    for cell in row:
                        if cell not in (None, ""):
                            vals.append(str(cell).strip())
    return sorted(set(vals))


def _parse_cell_ref(cell: str) -> Tuple[str, Optional[int]]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell.strip())
    if not m:
        return cell, None
    return m.group(1).upper(), int(m.group(2))


def _zip_filter_cell(path: str, sheet: str, cell: str) -> List[str]:
    col, row = _parse_cell_ref(cell)
    if row is None:
        return []
    # Try to read Data Validation for the specific cell via OOXML, including x14 extensions
    try:
        with zipfile.ZipFile(path, "r") as z:
            wb = z.read("xl/workbook.xml")
            import xml.etree.ElementTree as ET

            wb_root = ET.fromstring(wb)
            sheets_el = wb_root.find(
                ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets"
            )
            rId = None
            for sh in sheets_el:
                if sh.attrib.get("name") == sheet:
                    rId = sh.attrib.get(
                        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                    )
                    break
            if not rId:
                return []
            rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            ws_path = None
            for rel in rels_root:
                if rel.attrib.get("Id") == rId and rel.attrib.get("Type", "").endswith(
                    "/worksheet"
                ):
                    ws_path = "xl/" + rel.attrib.get("Target")
                    break
            if not ws_path:
                return []
            ws_root = ET.fromstring(z.read(ws_path))

            def col_to_idx(c: str) -> int:
                n = 0
                for ch in c:
                    n = n * 26 + (ord(ch.upper()) - 64)
                return n

            def cell_in_sqref(sqref_text: str) -> bool:
                if not sqref_text:
                    return False
                tgt = (col + str(row)).upper()
                parts = sqref_text.split()
                for p in parts:
                    if ":" in p:
                        a, b = p.split(":", 1)
                    else:
                        a = b = p
                    if a.upper() == tgt or b.upper() == tgt:
                        return True
                    # range check
                    m1 = re.match(r"\$?([A-Z]+)\$?(\d+)", a)
                    m2 = re.match(r"\$?([A-Z]+)\$?(\d+)", b)
                    if not (m1 and m2):
                        continue
                    ca, ra = m1.group(1), int(m1.group(2))
                    cb, rb = m2.group(1), int(m2.group(2))
                    ci = col_to_idx(col)
                    if col_to_idx(ca) <= ci <= col_to_idx(cb) and ra <= row <= rb:
                        return True
                return False

            formulas: List[str] = []
            # legacy dataValidations
            for dv in ws_root.findall(
                ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}dataValidation"
            ):
                if dv.attrib.get("type") == "list" and cell_in_sqref(
                    dv.attrib.get("sqref", "")
                ):
                    f1 = dv.find(
                        "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}formula1"
                    )
                    if f1 is not None and f1.text:
                        formulas.append(f1.text)

            # x14 extension dataValidations
            for ext in ws_root.findall(
                ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}extLst/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}ext"
            ):
                try:
                    x14ns = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
                    xmns = "http://schemas.microsoft.com/office/spreadsheetml/2006/main"
                    x = ET.fromstring(ET.tostring(ext))
                    for dv in x.findall(".//{" + x14ns + "}dataValidation"):
                        if dv.attrib.get("type") == "list":
                            sqref_el = dv.find(".//{" + xmns + "}sqref")
                            if not cell_in_sqref(
                                sqref_el.text if sqref_el is not None else ""
                            ):
                                continue
                            f1 = dv.find(".//{" + x14ns + "}formula1")
                            if f1 is not None:
                                f = f1.find(".//{" + xmns + "}f")
                                if f is not None and f.text:
                                    formulas.append(f.text)
                except Exception:
                    continue

            # Resolve formulas to values
            from openpyxl import load_workbook as _lb
            from openpyxl.utils import range_boundaries

            wb = _lb(path, data_only=True, read_only=True, keep_vba=True)
            vals: List[str] = []

            def read_range(ref: str) -> List[str]:
                ref2 = ref[1:] if ref.startswith("=") else ref
                # Named range or direct
                if "!" not in ref2:
                    if ref2 in wb.defined_names:
                        dn = wb.defined_names[ref2]
                        acc: List[str] = []
                        for title, area in dn.destinations:
                            ws2 = wb[title]
                            min_c, min_r, max_c, max_r = range_boundaries(area)
                            for rowv in ws2.iter_rows(
                                min_row=min_r,
                                max_row=max_r,
                                min_col=min_c,
                                max_col=max_c,
                                values_only=True,
                            ):
                                for cellv in rowv:
                                    if cellv not in (None, ""):
                                        acc.append(str(cellv).strip())
                        return acc
                    return []
                # sheet!range
                if ref2.startswith("'"):
                    sh = ref2.split("'", 2)[1]
                    rng = ref2.split("'", 2)[2].lstrip("!")
                else:
                    sh, rng = ref2.split("!", 1)
                ws2 = wb[sh]
                acc: List[str] = []
                min_c, min_r, max_c, max_r = range_boundaries(rng)
                for rowv in ws2.iter_rows(
                    min_row=min_r,
                    max_row=max_r,
                    min_col=min_c,
                    max_col=max_c,
                    values_only=True,
                ):
                    for cellv in rowv:
                        if cellv not in (None, ""):
                            acc.append(str(cellv).strip())
                return acc

            out: List[str] = []
            for f in formulas:
                ft = f.strip()
                # Direct quoted list: "A,B,C"
                if ft.startswith('"') and ft.endswith('"'):
                    out.extend([x.strip() for x in ft.strip('"').split(",") if x.strip()])
                else:
                    # Simple resolution; complex INDIRECTs will be handled by custom fallbacks
                    out.extend(read_range(ft))
            return sorted(set(out))
    except Exception:
        return []
    return []


def _supplier_tax_cell_named_range_fallback(path: str, sheet: str, cell: str) -> List[str]:
    # Workbook-specific heuristic: country-based named ranges in "Tenant Default Named Ranges"
    # For Supplier Tax, Column E depends on country in the same row (commonly at column C, or H)
    col, row = _parse_cell_ref(cell)
    if sheet != "Supplier Tax" or col != "E" or row is None:
        return []
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries

        wb = load_workbook(path, data_only=True, read_only=False, keep_vba=True)
        ws = wb[sheet]
        country = None
        # Try column C first, then H
        for cidx in (3, 8):
            v = ws.cell(row=row, column=cidx).value
            if isinstance(v, str) and v.strip():
                country = v.strip()
                break
        if not country:
            return []
        key = re.sub(r"[^A-Za-z0-9_]+", "_", country).strip("_") + "_Tax_Type"
        if key not in wb.defined_names:
            return []
        dn = wb.defined_names[key]
        vals: List[str] = []
        for title, area in dn.destinations:
            ws2 = wb[title]
            min_c, min_r, max_c, max_r = range_boundaries(area)
            for rowv in ws2.iter_rows(
                min_row=min_r,
                max_row=max_r,
                min_col=min_c,
                max_col=max_c,
                values_only=True,
            ):
                for cellv in rowv:
                    if cellv not in (None, ""):
                        vals.append(str(cellv).strip())
        return sorted(set(vals))
    except Exception:
        return []


def _distinct_values(path: str, sheet: str, col: str) -> List[str]:
    from openpyxl import load_workbook

    idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
    ws = load_workbook(path, data_only=True, read_only=True, keep_vba=True)[sheet]
    vals = set()
    for row in ws.iter_rows(min_col=idx, max_col=idx, values_only=True):
        v = row[0]
        if v not in (None, ""):
            vals.add(str(v).strip())
    return sorted(vals)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True)
    p.add_argument("--sheet", required=True)
    p.add_argument("--column", required=True)
    p.add_argument("--mode", choices=["auto", "filter", "distinct"], default="auto")
    p.add_argument("--cell")
    p.add_argument("--row", type=int)
    args = p.parse_args()
    path = os.path.expanduser(args.workbook)
    mode = args.mode.lower()
    vals: List[str] = []
    out_mode = "filter"
    if mode in ("auto", "filter"):
        # 1) Try cell-targeted xlwings validation first (if a cell was provided)
        vals = _try_xlwings_filter(path, args.sheet, args.column, args.cell, args.row)
        # 2) If nothing via xlwings and a cell is targeted, try OOXML parsing for that cell
        if not vals and args.cell:
            try:
                vals = _zip_filter_cell(path, args.sheet, args.cell)
            except Exception:
                vals = []
        # 3) Workbook-specific fallback for Supplier Tax → E[row] via named ranges by country
        if not vals and args.cell:
            sp_vals = _supplier_tax_cell_named_range_fallback(path, args.sheet, args.cell)
            if sp_vals:
                vals = sp_vals
        # 4) Column-wide validations as a fallback when no cell-specific values were found
        if not vals:
            try:
                vals = _openpyxl_filter(path, args.sheet, args.column)
            except Exception:
                vals = []
        if not vals and mode == "auto":
            vals = _distinct_values(path, args.sheet, args.column)
            out_mode = "distinct"
    elif mode == "distinct":
        vals = _distinct_values(path, args.sheet, args.column)
        out_mode = "distinct"
    print(
        json.dumps(
            {
                "sheet": args.sheet,
                "column": args.column,
                "mode": out_mode if vals else ("filter" if mode != "distinct" else "distinct"),
                "count": len(vals),
                "values": vals,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
