import argparse
import json
import os
from typing import Iterable, List, Optional


def _try_xlwings_filter(path: str, sheet: str, col: str, cell: Optional[str], row: Optional[int]) -> List[str]:
    try:
        import xlwings as xw
    except Exception:
        return []
    app = None
    vals = set()
    try:
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(path)
        sht = wb.sheets[sheet]
        targets: Iterable[int]
        target_rows: List[int] = []
        if cell:
            digits = "".join(ch for ch in cell if ch.isdigit())
            if digits.isdigit():
                target_rows.append(int(digits))
        elif row:
            target_rows.append(int(row))
        if not target_rows:
            used = sht.used_range
            last_row = used.last_cell.row
            end_row = min(last_row, 10000)
            target_rows = list(range(1, end_row + 1))
        targets = target_rows
        col_idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
        for r in targets:
            cell_rng = sht.range((r, col_idx))
            try:
                v = cell_rng.api.Validation
                t = v.Type
            except Exception:
                continue
            if t != 3:
                continue
            f = None
            try:
                f = v.Formula1
            except Exception:
                continue
            if not f:
                continue
            ref = f[1:] if isinstance(f, str) and f.startswith("=") else f
            if all(x not in ref for x in ("!", ":")):
                for part in str(ref).replace(";", ",").split(","):
                    p = part.strip()
                    if p:
                        vals.add(p)
                continue
            try:
                if "!" in ref:
                    sh, rng = ref.split("!", 1)
                    sh = sh.strip("'")
                    rng_obj = wb.sheets[sh].range(rng)
                else:
                    rng_obj = wb.names[ref].refers_to_range
                data = rng_obj.value
                if isinstance(data, list):
                    for rowv in data:
                        if isinstance(rowv, list):
                            for x in rowv:
                                if x not in (None, ""):
                                    vals.add(str(x).strip())
                        else:
                            if rowv not in (None, ""):
                                vals.add(str(rowv).strip())
                elif data not in (None, ""):
                    vals.add(str(data).strip())
            except Exception:
                continue
        if not vals:
            try:
                api = sht.api
                dds = api.DropDowns()
                for i in range(1, dds.Count + 1):
                    try:
                        dd = dds.Item(i)
                        tl = dd.TopLeftCell
                        col_match = tl.Column == (ord(col.upper()) - 64 if not col.isdigit() else int(col))
                        if not col_match:
                            continue
                        lfr = dd.ListFillRange
                        if not lfr:
                            continue
                        if "!" in lfr:
                            sh, rng = lfr.split("!", 1)
                            sh = sh.strip("'")
                            rng_obj = wb.sheets[sh].range(rng)
                        else:
                            rng_obj = wb.names[lfr].refers_to_range
                        data = rng_obj.value
                        if isinstance(data, list):
                            for rowv in data:
                                if isinstance(rowv, list):
                                    for x in rowv:
                                        if x not in (None, ""):
                                            vals.add(str(x).strip())
                                else:
                                    if rowv not in (None, ""):
                                        vals.add(str(rowv).strip())
                        elif data not in (None, ""):
                            vals.add(str(data).strip())
                    except Exception:
                        continue
            except Exception:
                pass
    finally:
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass
    return sorted(vals)


def _openpyxl_filter(path: str, sheet: str, col: str) -> List[str]:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries

    ws = load_workbook(path, data_only=True, keep_vba=True)[sheet]
    col_idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
    vals: List[str] = []
    dv_container = getattr(ws, "data_validations", None)
    if dv_container is not None:
        for dv in dv_container.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            applies = any(cr.min_col <= col_idx <= cr.max_col for cr in dv.ranges)
            if not applies:
                continue
            f1 = dv.formula1
            if not f1:
                continue
            f = f1[1:] if isinstance(f1, str) and f1.startswith("=") else f1
            if isinstance(f, str) and f.startswith('"') and f.endswith('"'):
                vals += [x.strip() for x in f.strip('"').split(",") if x.strip()]
                continue
            wb = ws.parent
            if isinstance(f, str) and "!" not in f and f in wb.defined_names:
                dn = wb.defined_names[f]
                for title, area in dn.destinations:
                    ws2 = wb[title]
                    min_c, min_r, max_c, max_r = range_boundaries(area)
                    for row in ws2.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
                        for cell in row:
                            if cell not in (None, ""):
                                vals.append(str(cell).strip())
                continue
            if isinstance(f, str) and "!" in f:
                if f.startswith("'"):
                    sh = f.split("'", 2)[1]
                    rng = f.split("'", 2)[2].lstrip("!")
                else:
                    sh, rng = f.split("!", 1)
                ws2 = wb[sh]
                min_c, min_r, max_c, max_r = range_boundaries(rng)
                for row in ws2.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True):
                    for cell in row:
                        if cell not in (None, ""):
                            vals.append(str(cell).strip())
    return sorted(set(vals))


def _distinct_values(path: str, sheet: str, col: str) -> List[str]:
    from openpyxl import load_workbook

    idx = ord(col.upper()) - 64 if not col.isdigit() else int(col)
    ws = load_workbook(path, data_only=True, read_only=True, keep_vba=True)[sheet]
    vals = set()
    for row in ws.iter_rows(min_col=idx, max_col=idx, values_only=True):
        v = row[0]
        if v not in (None, ""):
            vals.add(str(v).strip())
    return sorted(vals)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", required=True)
    p.add_argument("--sheet", required=True)
    p.add_argument("--column", required=True)
    p.add_argument("--mode", choices=["auto", "filter", "distinct"], default="auto")
    p.add_argument("--cell")
    p.add_argument("--row", type=int)
    args = p.parse_args()
    path = os.path.expanduser(args.workbook)
    mode = args.mode.lower()
    vals: List[str] = []
    out_mode = "filter"
    if mode in ("auto", "filter"):
        vals = _try_xlwings_filter(path, args.sheet, args.column, args.cell, args.row)
        if not vals:
            try:
                vals = _openpyxl_filter(path, args.sheet, args.column)
            except Exception:
                vals = []
        if not vals and mode == "auto":
            vals = _distinct_values(path, args.sheet, args.column)
            out_mode = "distinct"
    elif mode == "distinct":
        vals = _distinct_values(path, args.sheet, args.column)
        out_mode = "distinct"
    print(
        json.dumps(
            {
                "sheet": args.sheet,
                "column": args.column,
                "mode": out_mode if vals else ("filter" if mode != "distinct" else "distinct"),
                "count": len(vals),
                "values": vals,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
