
# examples/fin_workday_ref.py
"""Pipeline runner for Finance Workday Reference data."""

import json
from pathlib import Path
from base_runner import create_parser
import duckdb
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

if __name__ == "__main__":
    parser = create_parser("Run Finance Workday Reference pipeline")
    args = parser.parse_args()

    # Define paths
    config_base_path = Path(__file__).parent.parent / "config"
    publish_config_path = config_base_path / "publish/finance/publish_workday_ref.json"
    database_path = Path("~/Documents/__data/duckdb/rpatel.duckdb").expanduser()

    print("--- Starting DIRECT EXPORT (DuckDB -> Excel) ---")
    
    if not publish_config_path.exists():
        raise FileNotFoundError(f"Config not found: {publish_config_path}")
        
    # Load config manually
    with open(publish_config_path, "r") as f:
        config_data = json.load(f)
        
    # Connect to DB
    con = duckdb.connect(str(database_path), read_only=True)
    
    try:
        for wb_conf in config_data:
            tgt_dir = Path(wb_conf["tgtWorkbookPathName"]).expanduser()
            tgt_name = wb_conf["tgtWorkbookFileName"]
            tgt_path = tgt_dir / f"{tgt_name}.xlsx"
            
            print(f"Creating workbook: {tgt_path}")
            
            # Create new workbook
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
                
            sheets_conf = wb_conf.get("sheets", [])
            for sheet_conf in sheets_conf:
                sheet_name = sheet_conf["sheetName"]
                table_name = sheet_conf["srcTableName"]
                
                print(f"  - Exporting table '{table_name}' to sheet '{sheet_name}'...")
                
                try:
                    # Fetch data
                    df = con.execute(f"SELECT * FROM {table_name}").df()
                    
                    # Create sheet
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Write headers
                    headers = list(df.columns)
                    ws.append(headers)
                    
                    # Write data
                    for row in df.itertuples(index=False, name=None):
                        ws.append(row)
                        
                    # --- APPLY FORMATTING ---
                    
                    # 1. Header Styling (Bold, White Text, Blue Background)
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        
                    # 2. Auto-filter
                    ws.auto_filter.ref = ws.dimensions
                    
                    # 3. Freeze Top Row
                    ws.freeze_panes = "A2"
                    
                    # 4. Auto-size columns
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        # Iterate through all cells in column to find max width
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                                
                        # Set width (add a little padding)
                        adjusted_width = (max_length + 2)
                        # Cap max width to avoid excessively wide columns (e.g. 100 chars)
                        adjusted_width = min(adjusted_width, 100)
                        ws.column_dimensions[column_letter].width = adjusted_width

                except Exception as e:
                    print(f"    ERROR exporting table {table_name}: {e}")
            
            # Ensure target directory exists
            tgt_dir.mkdir(parents=True, exist_ok=True)
            
            # Save
            wb.save(tgt_path)
            print(f"Saved: {tgt_path}")

    finally:
        con.close()
